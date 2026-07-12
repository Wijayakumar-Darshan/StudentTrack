"""
excel_parser.py
Two Excel mark-sheet formats:
  Junior  – Grades 6-9  (e.g. First__team__Test_8D_2026.xlsm)
  Senior  – Grades 10-11 O/L (e.g. 2026__grade_11_First_Term___S1.xlsm)
"""
import re
import gc
from openpyxl import load_workbook

CLASS_SHEETS   = set("ABCDEFGH")
GRADE_ROW_IDX  = 2          # row index (0-based) that contains grade & class
GRADE_COL_JNR  = 6          # col index of grade number in Junior header
CLASS_COL_JNR  = 7          # col index of class letter in Junior header

# ── Junior subject column map (0-indexed) ───────────────────────────────────
JUNIOR_COLS = {
    2:  "Sinhala Language",
    3:  "Tamil Language",
    4:  "Buddhism",
    5:  "Shaivism",
    6:  "Catholic Doctrine",
    7:  "Christianity",
    8:  "English",
    9:  "Mathematics",
    10: "Science",
    11: "History",
    12: "Geography",
    13: "Life Skills",
    14: "Music (Western)",
    15: "Music (Oriental)",
    16: "Art",
    17: "Dance",
    18: "Drama",
    19: "ICT",
    20: "Practical & Technical Skills",
    21: "Health & Physical Education",
    22: "Second Language (Sinhala)",
    23: "Second Language (Tamil)",
}

# ── Senior O/L subject column map (LEFT half, 0-indexed) ────────────────────
OL_LEFT = {
    4:  "Sinhala Language",
    5:  "Tamil Language",
    6:  "Catholic Doctrine",
    7:  "Buddhism",
    10: "English",
    11: "Mathematics",
    12: "Science",
    13: "History",
    14: "Business & Accounting",
    15: "Civic Education",
    16: "Second Language (Tamil)",
    17: "Classical / Modern Languages",
    22: "Music (Oriental)",
    23: "Music (Western)",
    24: "Art",
    25: "Dance",
    26: "Literature (English)",
    27: "Literature (Sinhala)",
    28: "Drama & Performing Arts",
    30: "ICT",
    31: "Agriculture & Food Technology",
    32: "Health & Physical Education",
    33: "Media Studies",
}

# RIGHT-half offset: second student per row in Senior sheets
# Detected dynamically but default is 156
_OL_DEFAULT_OFFSET = 156


def _find_right_offset(sample_rows):
    """Find column offset where the second student starts in Senior sheets."""
    for row in sample_rows:
        for start in range(100, min(250, len(row))):
            try:
                if (isinstance(row[start], (int, float)) and
                        1 <= row[start] <= 60 and
                        start + 1 < len(row) and
                        isinstance(row[start + 1], str) and
                        len(row[start + 1].strip()) > 3):
                    return start
            except Exception:
                pass
    return _OL_DEFAULT_OFFSET


def _safe_mark(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        f = float(val)
        return round(f, 1) if 0 <= f <= 100 else None
    s = str(val).strip().lower()
    if s in ('ab', '-', '+', ''):
        return None
    try:
        f = float(s)
        return round(f, 1) if 0 <= f <= 100 else None
    except ValueError:
        return None


def _is_student_row(row, name_col=1):
    try:
        no   = row[0]
        name = row[name_col]
        return (isinstance(no, (int, float)) and int(no) == no and no > 0 and
                isinstance(name, str) and len(name.strip()) > 1)
    except Exception:
        return False


def _parse_marks(row, col_map):
    out = {}
    for ci, subj in col_map.items():
        if ci < len(row):
            v = _safe_mark(row[ci])
            if v is not None:
                out[subj] = v
    return out


def _grade_from_filename(path):
    name = str(path)
    m = re.search(r'grade[_\s]*(\d+)', name, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r'_(\d+)[A-Z]?_', name)
    if m:
        g = int(m.group(1))
        if 6 <= g <= 13:
            return g
    return None


def _year_term_from_rows(rows):
    year, term = None, 1
    for row in rows[:8]:
        for val in row:
            if isinstance(val, (int, float)) and 2020 <= val <= 2035:
                year = int(val)
            if isinstance(val, str):
                tl = val.lower()
                if 'first' in tl or 'පළමු' in tl:
                    term = 1
                elif 'second' in tl or 'දෙවන' in tl:
                    term = 2
                elif 'third' in tl or 'තෙවන' in tl:
                    term = 3
    return year, term


# ── Safe workbook handling (fixes Windows PermissionError) ───────────────────
def _load_workbook_safe(filepath):
    """Load workbook with read_only mode."""
    return load_workbook(filepath, read_only=True, data_only=True)


def _close_workbook(wb):
    """Force close and help garbage collection (important on Windows)."""
    if wb is not None:
        try:
            wb.close()
        except Exception:
            pass
        del wb
        gc.collect()


# ── Junior parser (grades 6-9) ───────────────────────────────────────────────

def parse_junior_excel(filepath, hint_grade=None):
    wb = None
    try:
        wb = _load_workbook_safe(filepath)
        results = []

        for sname in wb.sheetnames:
            if sname.strip().upper() not in CLASS_SHEETS:
                continue
            ws   = wb[sname]
            rows = list(ws.iter_rows(values_only=True))

            # Read grade & class from specific header row (row index 2)
            grade, cls = hint_grade, sname.strip().upper()
            if len(rows) > GRADE_ROW_IDX:
                hr = rows[GRADE_ROW_IDX]
                if (GRADE_COL_JNR < len(hr) and
                        isinstance(hr[GRADE_COL_JNR], (int, float)) and
                        6 <= hr[GRADE_COL_JNR] <= 13):
                    grade = int(hr[GRADE_COL_JNR])
                if (CLASS_COL_JNR < len(hr) and
                        isinstance(hr[CLASS_COL_JNR], str) and
                        hr[CLASS_COL_JNR].strip().upper() in CLASS_SHEETS):
                    cls = hr[CLASS_COL_JNR].strip().upper()

            year, term = _year_term_from_rows(rows)
            students   = []
            for row in rows:
                if not _is_student_row(row):
                    continue
                marks = _parse_marks(row, JUNIOR_COLS)
                if marks:
                    students.append({
                        "seq_no": int(row[0]),
                        "name":   str(row[1]).strip(),
                        "marks":  marks,
                    })
            if students:
                results.append({"class_section": cls, "grade": grade,
                                "term": term, "year": year, "students": students})
        return results
    finally:
        _close_workbook(wb)


# ── Senior O/L parser (grades 10-11) ─────────────────────────────────────────

def parse_senior_excel(filepath, hint_grade=None):
    wb = None
    try:
        wb = _load_workbook_safe(filepath)
        results = []

        for sname in wb.sheetnames:
            if sname.strip().upper() not in CLASS_SHEETS:
                continue
            ws   = wb[sname]
            rows = list(ws.iter_rows(values_only=True))

            # Grade from header row: look for '11' or '10' next to class letter
            grade, cls = hint_grade, sname.strip().upper()
            for row in rows[:8]:
                for ci, val in enumerate(row):
                    if isinstance(val, (int, float)) and 10 <= val <= 13:
                        # Check neighbour is a class letter
                        nxt = row[ci + 1] if ci + 1 < len(row) else None
                        if isinstance(nxt, str) and nxt.strip().upper() in CLASS_SHEETS:
                            grade = int(val)
                            cls   = nxt.strip().upper()
                            break

            year, term = _year_term_from_rows(rows)

            # Find right-half offset
            student_rows = [r for r in rows if _is_student_row(r)]
            offset = _find_right_offset(student_rows[:10]) if student_rows else _OL_DEFAULT_OFFSET

            # Build right-col map
            OL_RIGHT = {ci + offset: subj for ci, subj in OL_LEFT.items()}

            students = []
            seen     = set()

            def _add(no, name, marks):
                key = (int(no), name[:12])
                if key not in seen and marks:
                    seen.add(key)
                    students.append({"seq_no": int(no), "name": name, "marks": marks})

            for row in rows:
                if _is_student_row(row):
                    marks = _parse_marks(row, OL_LEFT)
                    _add(row[0], str(row[1]).strip(), marks)
                # Right half
                if len(row) > offset + 1:
                    rslice = row[offset:]
                    if _is_student_row(rslice):
                        marks = _parse_marks(row, OL_RIGHT)  # Note: using original row for indices
                        _add(rslice[0], str(rslice[1]).strip(), marks)

            if students:
                results.append({"class_section": cls, "grade": grade,
                                "term": term, "year": year, "students": students})
        return results
    finally:
        _close_workbook(wb)


# ── Auto-detect ───────────────────────────────────────────────────────────────

def detect_and_parse(filepath, hint_grade=None):
    wb = None
    try:
        if hint_grade is None:
            hint_grade = _grade_from_filename(filepath)

        wb = _load_workbook_safe(filepath)
        snames = wb.sheetnames
        # Check row width to distinguish formats
        senior = False
        for sn in snames:
            if sn.strip().upper() not in CLASS_SHEETS:
                continue
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                if _is_student_row(row) and len(row) > 100:
                    senior = True
                    break
            if senior:
                break
    finally:
        _close_workbook(wb)

    if senior or (hint_grade and hint_grade >= 10):
        return "senior", parse_senior_excel(filepath, hint_grade)
    return "junior", parse_junior_excel(filepath, hint_grade)